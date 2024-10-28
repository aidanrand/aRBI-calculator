from openpyxl import Workbook, load_workbook
import os

def write_results_to_excel(data, filename, regular_season):
    if excel_file_exists(filename):
        workbook = load_workbook(filename)

    else:
        # Write headers
        player_headers = ["Player Name", "Player ID", "March", "April", "May", "June", 
                   "July", "August", "September", "October", "November", "Total"]
        team_headers = ["Team Name", "March", "April", "May", "June", 
                   "July", "August", "September", "October", "November", "Total"]
        
        workbook = Workbook()
        sheet = workbook.create_sheet(title="Regular Season - Players")
        sheet.column_dimensions["A"].width = 25
        for col_num, header in enumerate(player_headers, 1):
            sheet.cell(row=1, column=col_num, value=header)
        
        sheet = workbook.create_sheet(title="Regular Season - Teams")
        sheet.column_dimensions["A"].width = 20
        for col_num, header in enumerate(team_headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        sheet = workbook.create_sheet(title="Postseason - Players")
        sheet.column_dimensions["A"].width = 25
        for col_num, header in enumerate(player_headers, 1):
            sheet.cell(row=1, column=col_num, value=header)
        
        sheet = workbook.create_sheet(title="Postseason - Teams")
        sheet.column_dimensions["A"].width = 20
        for col_num, header in enumerate(team_headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        workbook.remove(workbook.active)

    if regular_season:
        player_sheet = workbook["Regular Season - Players"]
        team_sheet = workbook["Regular Season - Teams"]
    else:
        player_sheet = workbook["Postseason - Players"]
        team_sheet = workbook["Postseason - Teams"]

    for player, id in data["home_players"].items():
        if data["home_aRBI"][player] == 0:
            continue
        player_found = False
        for row in range(2,player_sheet.max_row + 1):
            if player_sheet.cell(row=row, column=2).value == id:
                player_found = True
                old_value = player_sheet.cell(row=row, column=data["month"]).value
                new_value = data["home_aRBI"][player] + old_value
                player_sheet.cell(row=row, column=data["month"], value=new_value)
                break   
        if not player_found:
            player_sheet.append([player, id] + [0]*9)
            player_sheet.cell(row=player_sheet.max_row, column=data["month"], value=data["home_aRBI"][player])

    for player, id in data["away_players"].items():
        if data["away_aRBI"][player] == 0:
            continue
        player_found = False
        for row in range(2,player_sheet.max_row + 1):
            if player_sheet.cell(row=row, column=2).value == id:
                player_found = True
                old_value = player_sheet.cell(row=row, column=data["month"]).value
                new_value = data["away_aRBI"][player] + old_value
                player_sheet.cell(row=row, column=data["month"], value=new_value)
                break   
        if not player_found:
            player_sheet.append([player, id] + [0]*9)
            player_sheet.cell(row=player_sheet.max_row, column=data["month"], value=data["away_aRBI"][player])
    
    team_found = False
    for row in range(2, team_sheet.max_row + 1):
        if team_sheet.cell(row=row, column=1).value == data["home"]:
            team_found = True
            old_value = team_sheet.cell(row=row, column=data["month"]-1).value
            new_value = sum(data["home_aRBI"].values()) + old_value
            team_sheet.cell(row=row, column=data["month"]-1, value=new_value)
            break

    if not team_found:
        team_sheet.append([data["home"]] + [0]*9)
        team_sheet.cell(row=team_sheet.max_row, column=data["month"]-1, value=sum(data["home_aRBI"].values()))

    team_found = False
    for row in range(2, team_sheet.max_row + 1):
        if team_sheet.cell(row=row, column=1).value == data["away"]:
            team_found = True
            old_value = team_sheet.cell(row=row, column=data["month"]-1).value
            new_value = sum(data["away_aRBI"].values()) + old_value
            team_sheet.cell(row=row, column=data["month"]-1, value=new_value)
            break
    
    if not team_found:
        team_sheet.append([data["away"]] + [0]*9)
        team_sheet.cell(row=team_sheet.max_row, column=data["month"]-1, value=sum(data["away_aRBI"].values()))

    # Save the workbook
    workbook.save(filename)

def excel_file_exists(filename):
    return os.path.exists(filename)
